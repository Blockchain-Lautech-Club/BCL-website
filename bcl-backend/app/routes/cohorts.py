from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import io
from app.database import get_db
from app.db_models import CohortApplication, CohortSettings
from app.models import schemas as schemas
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
router = APIRouter()


# --- Helper: get or create the single settings row ---
def _get_settings(db: Session) -> CohortSettings:
    settings = db.query(CohortSettings).filter(CohortSettings.id == 1).first()
    if not settings:
        settings = CohortSettings(id=1, applications_open=True)
        db.add(settings)
        db.commit()
        db.refresh(settings)
    return settings


# --- Guard used inside register_cohort_member ---
@router.post("/applications", response_model=schemas.CohortResponse)
async def register_cohort_member(
    payload: schemas.CohortCreate,
    db: Session = Depends(get_db)
):
    # Block submissions when applications are closed
    settings = _get_settings(db)
    if not settings.applications_open:
        raise HTTPException(
            status_code=403,
            detail="Applications are currently closed."
        )

    existing_app = db.query(CohortApplication).filter(
        CohortApplication.email == payload.email
    ).first()
    if existing_app:
        raise HTTPException(
            status_code=400,
            detail="You have already submitted an application for this cohort."
        )

    new_application = CohortApplication(
        first_name=payload.first_name,
        last_name=payload.last_name,
        email=payload.email,
        github_handle=payload.github_handle,
        department=payload.department,
        level=payload.level,
        experience=payload.experience,
        registering_for=payload.registering_for,
        motivation=payload.motivation,
        commitment=payload.commitment,
        applied_at=datetime.now()
    )
    try:
        db.add(new_application)
        db.commit()
        db.refresh(new_application)
        return new_application
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail="Database error occurred.")

@router.get("/applications/export")
async def export_applications(db: Session = Depends(get_db)):
    applicants = (
        db.query(CohortApplication)
        .order_by(CohortApplication.applied_at.desc())
        .all()
    )
    if not applicants:
        raise HTTPException(status_code=404, detail="No applicants found.")

    wb = _build_excel(applicants)

    # Write to in-memory buffer — no disk writes on the server
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"cohort_applicants_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _build_excel(applicants):
    COLUMNS = [
        ("ID",              "id"),
        ("First Name",      "first_name"),
        ("Last Name",       "last_name"),
        ("Email",           "email"),
        ("GitHub Handle",   "github_handle"),
        ("Department",      "department"),
        ("Level",           "level"),
        ("Experience",      "experience"),
        ("Registering For", "registering_for"),
        ("Motivation",      "motivation"),
        ("Commitment",      "commitment"),
        ("Terms Agreed",    "terms"),
        ("Applied At",      "applied_at"),
    ]
    COL_WIDTHS = {
        "ID": 6, "First Name": 16, "Last Name": 16, "Email": 30,
        "GitHub Handle": 20, "Department": 18, "Level": 12,
        "Experience": 14, "Registering For": 18, "Motivation": 40,
        "Commitment": 13, "Terms Agreed": 13, "Applied At": 22,
    }
    HEADER_FILL = PatternFill("solid", start_color="1F3864")
    ALT_FILL    = PatternFill("solid", start_color="EEF2FF")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    CELL_FONT   = Font(name="Arial", size=10)
    BORDER      = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),  bottom=Side(style="thin", color="D0D0D0"),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Applicants"
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, (header, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 15)
    ws.row_dimensions[1].height = 28

    # Data rows
    for row_idx, applicant in enumerate(applicants, start=2):
        for col_idx, (_, field) in enumerate(COLUMNS, start=1):
            value = getattr(applicant, field, None)
            if isinstance(value, bool):
                value = "Yes" if value else "No"
            elif isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = CELL_FONT
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(field == "motivation"))
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # Summary sheet
    summary = wb.create_sheet("Summary")
    summary.merge_cells("A1:B1")
    summary["A1"].value = "Cohort Applications — Summary"
    summary["A1"].font  = Font(name="Arial", bold=True, size=14, color="1F3864")

    total     = len(applicants)
    committed = sum(1 for a in applicants if a.commitment)
    by_dept   = {}
    by_track  = {}
    for a in applicants:
        by_dept[a.department]       = by_dept.get(a.department, 0) + 1
        by_track[a.registering_for] = by_track.get(a.registering_for, 0) + 1

    rows = [
        [], ["Metric", "Value"],
        ["Total Applicants", total],
        ["Committed (Yes)",  committed],
        [], ["— By Department —", ""],
        *[(d, c) for d, c in sorted(by_dept.items())],
        [], ["— By Track —", ""],
        *[(t, c) for t, c in sorted(by_track.items())],
    ]
    for row in rows:
        summary.append(row)
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 16

    return wb

# --- Admin: open applications ---
@router.patch("/applications/open", summary="Open cohort applications")
async def open_applications(db: Session = Depends(get_db)):
    settings = _get_settings(db)
    if settings.applications_open:
        return {"message": "Applications are already open.", "applications_open": True}
    
    settings.applications_open = True
    settings.updated_at = datetime.now()
    db.commit()
    db.refresh(settings)
    return {"message": "Applications are now open.", "applications_open": True}


# --- Admin: close applications ---
@router.patch("/applications/close", summary="Close cohort applications")
async def close_applications(db: Session = Depends(get_db)):
    settings = _get_settings(db)
    if not settings.applications_open:
        return {"message": "Applications are already closed.", "applications_open": False}
    
    settings.applications_open = False
    settings.updated_at = datetime.now()
    db.commit()
    db.refresh(settings)
    return {"message": "Applications are now closed.", "applications_open": False}


# --- Optional: check current status (useful for frontend banners) ---
@router.get("/applications/status", summary="Check if applications are open")
async def application_status(db: Session = Depends(get_db)):
    settings = _get_settings(db)
    return {
        "applications_open": settings.applications_open,
        "updated_at": settings.updated_at
    }