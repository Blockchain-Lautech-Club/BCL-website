"""
export_applicants.py
--------------------
Exports cohort applicants from the database to a formatted Excel file.

Usage:
    python export_applicants.py                        # exports all applicants
    python export_applicants.py --output my_file.xlsx  # custom output path
"""

import argparse
from datetime import datetime
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import SQLALCHEMY_DATABASE_URL
from app.models import CohortApplication


# ── DB setup ────────────────────────────────────────────────────────────────

def get_session():
    engine = create_engine(SQLALCHEMY_DATABASE_URL)
    Session = sessionmaker(bind=engine)
    return Session()


def fetch_applicants(db):
    return (
        db.query(CohortApplication)
        .order_by(CohortApplication.applied_at.desc())
        .all()
    )


# ── Excel builder ────────────────────────────────────────────────────────────

COLUMNS = [
    ("ID",              "id"),
    ("First Name",      "first_name"),
    ("Last Name",       "last_name"),
    ("Email",           "email"),
    ("GitHub Handle",   "github_handle"),
    ("Department",      "department"),
    ("Level",           "level"),
    ("Experience",      "experience"),
    ("Registering For", "registering_for"),
    ("Motivation",      "motivation"),
    ("Commitment",      "commitment"),
    ("Terms Agreed",    "terms"),
    ("Applied At",      "applied_at"),
]

HEADER_FILL   = PatternFill("solid", start_color="1F3864")   # dark navy
ALT_ROW_FILL  = PatternFill("solid", start_color="EEF2FF")   # soft blue-white
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CELL_FONT     = Font(name="Arial", size=10)
THIN_BORDER   = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

COL_WIDTHS = {
    "ID": 6, "First Name": 16, "Last Name": 16, "Email": 30,
    "GitHub Handle": 20, "Department": 18, "Level": 12,
    "Experience": 14, "Registering For": 18, "Motivation": 40,
    "Commitment": 13, "Terms Agreed": 13, "Applied At": 22,
}


def build_excel(applicants: list, output_path: str):
    wb = Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────────
    summary = wb.active
    summary.title = "Summary"

    summary.merge_cells("A1:C1")
    title_cell = summary["A1"]
    title_cell.value = "Cohort Applications — Summary"
    title_cell.font = Font(name="Arial", bold=True, size=14, color="1F3864")
    title_cell.alignment = Alignment(horizontal="left")

    summary.append([])
    summary.append(["Metric", "Value"])
    for cell in summary[3]:
        cell.font = Font(name="Arial", bold=True, size=11)

    total = len(applicants)
    committed = sum(1 for a in applicants if a.commitment)
    by_dept: dict = {}
    by_level: dict = {}
    by_track: dict = {}

    for a in applicants:
        by_dept[a.department]        = by_dept.get(a.department, 0) + 1
        by_level[a.level]            = by_level.get(a.level, 0) + 1
        by_track[a.registering_for]  = by_track.get(a.registering_for, 0) + 1

    summary_rows = [
        ("Total Applicants",     total),
        ("Committed (Yes)",      committed),
        ("Committed (%)",        f"=B5/B4*100" if total else "N/A"),
        ("",                     ""),
        ("--- By Department ---", ""),
        *[(dept, count) for dept, count in sorted(by_dept.items())],
        ("",                     ""),
        ("--- By Level ---",     ""),
        *[(lvl, count) for lvl, count in sorted(by_level.items())],
        ("",                     ""),
        ("--- By Track ---",     ""),
        *[(track, count) for track, count in sorted(by_track.items())],
    ]

    for row in summary_rows:
        summary.append(list(row))

    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 16

    # ── Applicants sheet ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Applicants")

    headers = [col[0] for col in COLUMNS]
    ws.append(headers)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 15)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for row_idx, applicant in enumerate(applicants, start=2):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else None
        for col_idx, (_, field) in enumerate(COLUMNS, start=1):
            value = getattr(applicant, field, None)

            # Normalise booleans and timestamps for readability
            if isinstance(value, bool):
                value = "Yes" if value else "No"
            elif isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M")

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = CELL_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(field == "motivation"))
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 18

    # Totals row
    total_row = len(applicants) + 2
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).font  = Font(name="Arial", bold=True, size=10)
    ws.cell(row=total_row, column=2).value = f"=COUNTA(B2:B{total_row - 1})"
    ws.cell(row=total_row, column=2).font  = Font(name="Arial", bold=True, size=10)

    wb.save(output_path)
    print(f"✅  Exported {total} applicant(s) → {output_path}")


# ── Entry point ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Export cohort applicants to Excel.")
    parser.add_argument(
        "--output", "-o",
        default=f"cohort_applicants_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        help="Output .xlsx file path (default: timestamped filename)"
    )
    args = parser.parse_args()

    db = get_session()
    try:
        applicants = fetch_applicants(db)
        if not applicants:
            print("⚠️  No applicants found in the database.")
            return
        build_excel(applicants, args.output)
    finally:
        db.close()


if __name__ == "__main__":
    main()